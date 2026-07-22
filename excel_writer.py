"""Gravação dos dados de saldo FAPEU na planilha Excel."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "saldo fapeu"
HEADERS = (
    "cdProjeto",
    "PRJ",
    "TotalReceita",
    "TotalDespesas",
    "TotalCLT",
    "TotalPessoalNaoContratado",
    "TotalRedoa",
)
COLUNAS_NUMERO = range(3, len(HEADERS) + 1)  # colunas C a G (TotalReceita..TotalRedoa)

SHEET_NAME_CLT_RPA = "saldo clt_rpa"
HEADERS_CLT_RPA = (
    "cdProjeto",
    "Ano",
    "Mes",
    "TotalCLT",
    "TotalPessoalNaoContratado",
)
COLUNAS_NUMERO_CLT_RPA = (4, 5)  # colunas D e E (TotalCLT, TotalPessoalNaoContratado)

# Formato numérico simples (sem símbolo de moeda e sem estilo contábil).
FORMATO_NUMERO = '#,##0.00'


class AvisoPlanilha(RuntimeError):
    """Levantado quando o arquivo ou a aba de destino não são encontrados."""


def _parse_valor_brl(valor: str | None) -> float:
    """Converte números formatados em pt-BR (ex.: '1.234,56') para float."""
    if valor is None:
        return 0.0
    texto = str(valor).strip().replace(".", "").replace(",", ".")
    return float(texto) if texto else 0.0


def _escrever_aba(planilha, headers: tuple[str, ...], linhas: list[tuple], colunas_numero) -> None:
    """Limpa a aba e regrava com os dados atuais, convertendo as colunas numéricas de pt-BR para float."""
    planilha.delete_rows(1, planilha.max_row)
    planilha.append(headers)
    for linha in linhas:
        linha_num = planilha.max_row + 1
        valores = list(linha)
        for coluna in colunas_numero:
            valores[coluna - 1] = _parse_valor_brl(valores[coluna - 1])
        planilha.append(valores)
        for coluna in colunas_numero:
            planilha.cell(row=linha_num, column=coluna).number_format = FORMATO_NUMERO


def atualizar_planilha(
    caminho_planilha: Path,
    linhas: list[tuple],
    linhas_clt_rpa: list[tuple] | None = None,
) -> None:
    """Regrava a aba saldo fapeu e, se informado, a aba saldo clt_rpa com os dados atuais das consultas."""
    if not caminho_planilha.exists():
        raise AvisoPlanilha(
            f"Arquivo '{caminho_planilha.name}' não encontrado em:\n{caminho_planilha.parent}"
        )

    try:
        workbook = load_workbook(caminho_planilha)
    except PermissionError as exc:
        raise RuntimeError(
            f"Não foi possível abrir '{caminho_planilha.name}'. "
            "Feche o arquivo no Excel e tente novamente."
        ) from exc

    if SHEET_NAME not in workbook.sheetnames:
        raise AvisoPlanilha(f"Aba '{SHEET_NAME}' não encontrada em {caminho_planilha.name}")
    _escrever_aba(workbook[SHEET_NAME], HEADERS, linhas, COLUNAS_NUMERO)

    if linhas_clt_rpa is not None:
        if SHEET_NAME_CLT_RPA not in workbook.sheetnames:
            raise AvisoPlanilha(f"Aba '{SHEET_NAME_CLT_RPA}' não encontrada em {caminho_planilha.name}")
        _escrever_aba(workbook[SHEET_NAME_CLT_RPA], HEADERS_CLT_RPA, linhas_clt_rpa, COLUNAS_NUMERO_CLT_RPA)

    try:
        workbook.save(caminho_planilha)
    except PermissionError as exc:
        raise RuntimeError(
            f"Não foi possível salvar '{caminho_planilha.name}'. "
            "Feche o arquivo no Excel e tente novamente."
        ) from exc
